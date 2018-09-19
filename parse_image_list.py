#!/usr/bin/env python
# -*- coding:utf-8 -*-
__author__ = 'wangrong'
import zipfile
import os,re,shutil
import json
from openpyxl import Workbook
from openpyxl.styles import Font,PatternFill

## https://automatetheboringstuff.com/#toc

class ZFile():
    def __init__(self,src_zip_path):
        self.app_map={'Application':{},'Application_Stateful':{}}
        # self.exact_tmp_dir=os.path.join(os.path.dirname(os.path.abspath(__file__)),'tmp')
        # if os.path.exists(self.exact_tmp_dir):
        #     shutil.rmtree(self.exact_tmp_dir)
        #     os.makedirs(self.exact_tmp_dir)
        # else:
        #     os.makedirs(self.exact_tmp_dir)

        src_zip_path=src_zip_path.replace('\\','/')
        if src_zip_path[-1] == '/':
            self.src_path = src_zip_path[:-1]
        else:
            self.src_path = src_zip_path

    def is_correct_zip_file(self,zipinstance):
        res=zipinstance.testzip()
        if res is None:
            return True
        else:
            return False

    def parse_application_json(self,json_map):
        try:
            application_name = json_map['applicationName']
            image = json_map['appInfo']['images'][0]['image']
            self.app_map['Application'][application_name]=image
        except KeyError as e:
            return False
        return True

    def parse_application_stateful_json(self,json_map):
        try:
            application_name = json_map['applicationName']
            image = json_map['imageName']
            self.app_map['Application_Stateful'][application_name]=image
        except KeyError as e:
            return False
        return True

    def parse_json_file(self,json_file):
        json_dict=json.loads(json_file)
        app_res_status=self.parse_application_json(json_dict)
        if not app_res_status:
            stat_app_res_status=self.parse_application_stateful_json(json_dict)
            if not stat_app_res_status:
                return False
        return True

    def unzip_zip_file(self,zip_file,match_file='app.json',mode='r'):
        zfile = zipfile.ZipFile(zip_file, mode, compression=zipfile.ZIP_DEFLATED)
        res_code=self.is_correct_zip_file(zfile)
        if not res_code:
            print ("Zip File is error")
            exit()
        all_file_list=zfile.namelist()
        i_count=0
        for one in all_file_list:
            if re.search(match_file.lower(),one.lower()):
                i_count=i_count+1
                #zfile.extract(one,self.exact_tmp_dir)
                json_file=zfile.open(one,'r').read()
                #print (str(json_file,encoding='utf-8'))
                prese_status=self.parse_json_file(str(json_file,encoding='utf-8'))
                if not prese_status:
                    print ("Parse ZIP File:{} Failed".format(zip_file))
        if i_count <=0 :
            print ("{} file not found in ZIP FILE {}".format(match_file,zip_file))
        #print (self.app_map)

    def write_json_to_excel(self):
        wb = Workbook()
        sheet = wb.active
        sheet.title="镜像列表"
        fontObj1 = Font(name='宋体', bold=True,size=12)
        fill = PatternFill(fill_type='solid',fgColor="98FB98")
        if self.app_map['Application']:
            rownum = 2
            sheet['A1'].value="无状态应用名称"
            sheet['B1'].value="镜像地址"
            # sheet.freeze_panes='A1'
            # sheet.freeze_panes = 'B1'
            sheet['A1'].fill=fill
            sheet['B1'].fill = fill
            sheet['A1'].font =fontObj1
            sheet['B1'].font = fontObj1
            sheet.column_dimensions['A'].width = 30
            sheet.column_dimensions['B'].width = 70
            for one in self.app_map['Application']:
                sheet['A%d'%rownum].value=one
                sheet['B%d' % rownum].value = self.app_map['Application'][one]
                rownum+=1
        if self.app_map['Application_Stateful']:
            rownum = 2
            sheet['D1'].value = "有状态应用名称"
            sheet['E1'].value = "镜像地址"
            # sheet.freeze_panes = 'D1'
            # sheet.freeze_panes = 'E1'
            sheet['D1'].fill = fill
            sheet['E1'].fill = fill
            sheet['D1'].font = fontObj1
            sheet['E1'].font = fontObj1
            sheet.column_dimensions['D'].width = 30
            sheet.column_dimensions['E'].width = 70
            for one in self.app_map['Application_Stateful']:
                sheet['D%d'%rownum].value=one
                sheet['E%d' % rownum].value = self.app_map['Application_Stateful'][one]
                rownum+=1

        try:
            wb.save(self.src_path+'/'+'image.xlsx')
            print("Excel File Generate in {}".format(self.src_path + '/' + 'image.xlsx'))
        except PermissionError:
            wb.save(self.src_path + '/' + 'image_new.xlsx')
            print("Excel File Generate in {}".format(self.src_path + '/' + 'image_new.xlsx'))




    def main_process(self,skip_file=['sql.zip'], match_postfix=['.zip']):
        parse_app_file = []
        skip_zip_file = []
        for dirpath, dirnames, filenames in os.walk(self.src_path):
            for file in filenames:
                if os.path.splitext(file)[1] in match_postfix:
                    if skip_file:
                        for skip_key in skip_file:
                            if not re.search(skip_key, file.lower()):
                                fullfile = (dirpath + '/' + file).replace('\\', '/')
                                self.unzip_zip_file(fullfile)
                                parse_app_file.append(fullfile)
                            else:
                                fullfile = (dirpath + '/' + file).replace('\\', '/')
                                skip_zip_file.append(fullfile)
                    else:
                        fullfile = (dirpath + '/' + file).replace('\\', '/')
                        self.unzip_zip_file(fullfile)
                        parse_app_file.append(fullfile)

        #print (json.dumps(self.app_map,indent=1))
        self.write_json_to_excel()




if __name__ == '__main__':
    a=ZFile(r'D:\WORKHOME\格鲁')
    a.main_process()
    #a.unzip_zip_file(r'D:\WORKHOME\格鲁\irc-bpm-product.zip')
    #a.unzip_zip_file(r'D:\WORKHOME\格鲁\med-backend-product.zip')