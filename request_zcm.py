#!/usr/bin/env python
# -*- coding:utf-8 -*-
__author__ = 'wangrong'
import requests
import json
import datetime
import os,re
from openpyxl import Workbook
from openpyxl.styles import Font,PatternFill

project_code_map = {'dbeptest1': [562023, 3061],
                    'dbeptest2': [562024, 3061],
                    'dbepdc': [562057, 2],
                    'dbepuat1': [562063, 2]}
class ZcmRequest():

    def __init__(self, request_ip, port=None, IsHttps='N'):
        self.global_header = {'Content-Type': 'application/json;charset=UTF-8',
                              "Accept": "*/*"}
        if port is not None:
            if IsHttps.upper() == 'Y':
                self.request_url = 'https://' + request_ip + ':' + port
            else:
                self.request_url = 'http://' + request_ip + ':' + port
        else:
            if IsHttps.upper() == 'Y':
                self.request_url = 'https://' + request_ip
            else:
                self.request_url = 'http://' + request_ip
        self.postfix=datetime.datetime.now().strftime('%H%M%S')
        self.log_path = './log/{}'.format(datetime.datetime.now().strftime('%Y%m%d-%H'))
        if not os.path.exists(self.log_path):
            os.makedirs(self.log_path)



    def ExecuteSqlZipFromServer(self, fileName, filePath, ftpHost, userName, passWord,projectCode, ftpPort=22):
        payload = {}
        payload['fileName'] = fileName
        payload['filePath'] = filePath
        payload['ftpHost'] = ftpHost
        payload['userName'] = userName
        payload['passWord'] = passWord
        payload['ftpPort'] = ftpPort
        payload['projectCode'] = projectCode
        requests_url = self.request_url + '/DeployManage/deploy'
        r = requests.post(requests_url, headers=self.global_header,
                          data=json.dumps(payload))
        print("\033[1;32msscInstanceId ==>\033[0m  {} ".format(r.json()['data']['sscInstanceId']))

    def queryJobExecStatus(self, instanceId):
        payload = {"execType": 1}
        payload['instanceId'] = instanceId
        requests_url = self.request_url + '/jobManage/queryJobExecStatus'
        log_file = './log/{}'
        response = []
        r = requests.post(requests_url, headers=self.global_header,
                          data=json.dumps(payload))
        # print (r.json())
        if r.json()['status'] == 'running':
            print("\033[1;33mJob {} is running...\033[0m".format(instanceId))
            # print (json.dumps(r.json(), indent=1))
        elif r.json()['status'] == 'fail':
            # print (r.json()['data'])
            total_files = 0
            total_sql_rows = 0
            for one_row in r.json()['data']:
                temp_response = {}
                temp_response['url'] = one_row['url']
                temp_response['script'] = one_row['script']
                temp_response['status'] = one_row['status']
                temp_response['success'] = one_row['success']
                temp_response['fail'] = one_row['fail']
                temp_response['executed'] = one_row['executed']
                temp_response['total'] = one_row['total']
                temp_response['sqlDetail'] = one_row['sqlDetail']
                response.append(temp_response)
                total_files = total_files + 1
                total_sql_rows = total_sql_rows + one_row['total']
            errlog = '{}/err_{}.log'.format(self.log_path, instanceId)
            print("Execute error, log in {},Execute total < {} > files,Execute total < {} > sql rows".format(errlog,
                                                                                                             total_files,
                                                                                                             total_sql_rows))
            json.dump(response, open(errlog, "w"), indent=1)
        elif r.json()['status'] == 'success':
            total_files = 0
            total_sql_rows = 0
            for one_row in r.json()['data']:
                total_files = total_files + 1
                total_sql_rows = total_sql_rows + one_row['total']
            oklog = '{}/success_{}.log'.format(self.log_path, instanceId)
            json.dump(r.json(), open(oklog, "w"), indent=1)
            print(
                'Execute successfully,log in {},Execute total files is < {} > files,Execute total < {} > sql rows'.format(
                    oklog, total_files, total_sql_rows))
        else:
            print(json.dumps(r.json(), indent=1))

    def checkFile(self, filePath, fileName, ftpHost='172.16.17.21', userName='appserver', passWord='appserver',
                  ftpPort=22, scriptType='oracle'):
        payload = {}
        payload['filePath'] = filePath
        payload['fileName'] = fileName
        payload['ftpHost'] = ftpHost
        payload['userName'] = userName
        payload['passWord'] = passWord
        payload['ftpPort'] = ftpPort
        payload['scriptType'] = scriptType
        requests_url = self.request_url + '/parser/checkFile'
        r = requests.post(requests_url, headers=self.global_header, params=payload)
        if r.json()['data'] == 'parse success!':
            print('Parse Result:\033[1;32m{}\033[0m'.format('Pass'))
        else:
            print('Parse Result:\033[1;31m{},\n{}\033[0m'.format(r.json()['data'], r.json()['msg']))

    def write_map_to_excel(self,data_map):
        status=True
        wb = Workbook()
        sheet = wb.active
        sheet.title = "镜像列表"
        fontObj1 = Font(name='宋体', bold=True, size=12)
        fill = PatternFill(fill_type='solid', fgColor="98FB98")
        try:
            if data_map['Application']:
                rownum = 2
                sheet['A1'].value = "无状态应用名称"
                sheet['B1'].value = "镜像地址"
                # sheet.freeze_panes='A1'
                # sheet.freeze_panes = 'B1'
                sheet['A1'].fill = fill
                sheet['B1'].fill = fill
                sheet['A1'].font = fontObj1
                sheet['B1'].font = fontObj1
                sheet.column_dimensions['A'].width = 35
                sheet.column_dimensions['B'].width = 80
                for one in data_map['Application']:
                    sheet['A%d' % rownum].value = one
                    sheet['B%d' % rownum].value = data_map['Application'][one]
                    rownum += 1
            else:
                status = False
        except KeyError:
            status = False
            pass
        try:
            if data_map['Application_Stateful']:
                rownum = 2
                sheet['D1'].value = "有状态应用名称"
                sheet['E1'].value = "镜像地址"
                # sheet.freeze_panes = 'D1'
                # sheet.freeze_panes = 'E1'
                sheet['D1'].fill = fill
                sheet['E1'].fill = fill
                sheet['D1'].font = fontObj1
                sheet['E1'].font = fontObj1
                sheet.column_dimensions['D'].width = 35
                sheet.column_dimensions['E'].width = 80
                for one in data_map['Application_Stateful']:
                    sheet['D%d'%rownum].value=one
                    sheet['E%d' % rownum].value = data_map['Application_Stateful'][one]
                    rownum+=1
            else:
                status = False
        except KeyError:
            status = False
            pass

        if status:
            try:
                wb.save(self.log_path+'/'+'image.xlsx')
                print("Excel File Generate in {}".format(self.log_path + '/' + 'image.xlsx'))
            except PermissionError:
                wb.save(self.log_path + '/' + 'image_new.xlsx')
                print("Excel File Generate in {}".format(self.log_path + '/' + 'image_{}.xlsx'.format(self.postfix)))
        else:
            print("Application Not Found!")


    def get_applications(self, project_code,page=0,size=9999):
        global project_code_map
        payload = {}
        res = {'Application':{}}
        try:
            payload['projectId'] = project_code_map[project_code][0]
            payload['tenantId'] = project_code_map[project_code][1]
            payload['page']=page
            payload['size']=size
            app_url = self.request_url + '/portal/zcm-application/applications/search'
        except Exception as err:
            print("Not Support Project:{}".format(project_code))
            exit()
        r = requests.get(app_url, params=payload,timeout=10)
        json_map = r.json()
        if json_map:
            for one in json_map:
                applicationName = one['appBaseInfo']['applicationName']
                image = one['appBaseInfo']['image']
                res['Application'][applicationName] = image
        return res
        # self.write_map_to_excel(res)
        # print(json.dumps(res, indent=1))

    def get_stateful_application(self,project_code):
        global project_code_map
        payload={}
        res = {'Application_Stateful': {}}
        try:
            payload['projectId'] = project_code_map[project_code][0]
            payload['tenantId'] = project_code_map[project_code][1]
            app_url = self.request_url + '/portal/zcm-resource/StatefulApplication/select'
        except Exception as err:
            print ("Not Support Project:{}".format(project_code))
            exit()
        r = requests.get(app_url, params=payload, timeout=10)
        json_map = r.json()
        if json_map:
            for one in json_map:
                applicationName = one['applicationName']
                image = one['imageName']
                res['Application_Stateful'][applicationName] = image
        return res

    def get_app_image_list(self,project_code):
        res_app=self.get_applications(project_code)
        res_state_app=self.get_stateful_application(project_code)
        total_res=dict(res_app,**res_state_app)
        #print (total_res)
        self.write_map_to_excel(total_res)
        #print(json.dumps(total_res, indent=1))

if __name__ == '__main__':
    ssc = ZcmRequest('10.45.80.26', '18280')
    ###查询结果
    ssc.queryJobExecStatus(1100)
    ###执行sql.zip
    ### fileName, filePath, ftpHost, userName, passWord
    # ssc.ExecuteSqlZipFromServer("string","string","string","string","string")
    #################################################################################
    ###########批量执行sql###################################################
    ###########################################
    # product_name=['sett']
    # script_home='/lvdata/appserver/kunpeng42'
    # for product in product_name:
    #     filePath=script_home+'/'+product+'/sql'
    #     ssc.ExecuteSqlZipFromServer("sql.zip", filePath, "172.16.17.21", "appserver", "appserver")

    #################################################################################
    ###checkfile(default ftpHost='172.16.17.21',userName='appserver',passWord='appserver')#
    #################################################################################
    # ssc.checkFile('/lvdata/appserver/kunpeng/balc/sql','balc_balc_bc_oracle.sql')

    #################################################################################
    ####################checkfile in cycle##########################################
    #################################################################################
    # realpath_dir = r'C:/Users/cc/Desktop/解析sql'
    # prefix_path='/lvdata/appserver/parsesql'
    # match_postfix = ['.sql']
    # skip_keyword = ['qmdb']
    # sql_file_list = []
    # for dirpath, dirnames, filenames in os.walk(realpath_dir):
    #     fpath = dirpath.replace(realpath_dir, '')
    #     for filename in filenames:
    #         if os.path.splitext(filename)[1] in match_postfix:
    #             if skip_keyword:
    #                 for skip_key in skip_keyword:
    #                     if not re.search(skip_key, filename.lower()):
    #                         print ("checkFile \033[1;33m{}\033[0m ...".format(filename))
    #                         ssc.checkFile(prefix_path, filename)
    #                         #sql_file_list.append(filename)
    #                         # print (prefix_path+'/'+filename)
    #             else:
    #                 print("checkFile \033[1;33m{}\033[0m  ...".format(filename))
    #                 ssc.checkFile(prefix_path, filename)
